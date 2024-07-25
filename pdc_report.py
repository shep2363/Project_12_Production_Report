import openpyxl
import os
from xml.etree import ElementTree as ET
from datetime import datetime, time, timedelta


WORKING_TIME = 24

def time_to_seconds(timestr):
    """Convert time string formatted as H:M:S to seconds."""
    h, m, s = map(int, timestr.split(':'))
    return h * 3600 + m * 60 + s

def seconds_to_decimal_hours(seconds):
    """Convert seconds to a decimal representing the number of hours."""
    return seconds / 3600

def time_difference(time1, time2):
    """Calculate the difference between two time strings formatted as H:M:S and return the result in seconds."""
    seconds1 = time_to_seconds(time1)
    seconds2 = time_to_seconds(time2)
    difference = seconds2 - seconds1
    if difference < 0:
        difference += 24 * 60 * 60  # Adjust for time that goes over midnight
    return difference

def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Adjust the width
        worksheet.column_dimensions[column].width = adjusted_width

def get_shift_time():
    """Prompt the user to input the start and end times for day and night shifts."""
    day_shift_start = '07:00:00'
    day_shift_end = '19:00:00'
    night_shift_start = '19:01:00'
    night_shift_end = '06:59:00'
    
    return day_shift_start, day_shift_end, night_shift_start, night_shift_end

def determine_shift(start_time, day_shift_start, day_shift_end, night_shift_start, night_shift_end):
    """Determine if a time falls within the day or night shift."""
    start_seconds = time_to_seconds(start_time)
    day_start_seconds = time_to_seconds(day_shift_start)
    day_end_seconds = time_to_seconds(day_shift_end)
    night_start_seconds = time_to_seconds(night_shift_start)
    night_end_seconds = time_to_seconds(night_shift_end)

    # Assuming that shifts do not overlap. The logic here may need to be adjusted based on how shifts are scheduled.
    if day_start_seconds <= start_seconds < day_end_seconds:
        return 'Day'
    elif night_start_seconds <= start_seconds or start_seconds < night_end_seconds:  # Covering overnight shift
        return 'Night'
    else:
        return 'Undefined'  # For times that do not fall into either shift

# Get the XML file path from the user
xml_file_path = input("Enter the path to the XML file: ").strip('"')

# Parse the XML from file
tree = ET.parse(xml_file_path)
root = tree.getroot()

# Create a new Excel workbook and a master sheet
wb = openpyxl.Workbook()

# Step 1: Create the Dashboard sheet as the first sheet
dashboard = wb.create_sheet(title="Dashboard", index=0)  # index=0 makes it the first sheet

# Step 2: Write the headers to the Dashboard
dashboard.append(["Date", "Production Efficiency"])  # Add your headers here

# Dictionary to keep track of unique dates (to avoid duplicate entries in the Dashboard)
unique_dates = {}

master_sheet = wb.create_sheet(title="Master Sheet")
master_sheet.append(["Part Name", "Date", "Start Time", "Finish Time", "Total Run Time (hours)", "Idle Time (hours)", 
                     "Production Time (hours)", "PT-TRT (hours)", "Shift"])  # The header for the master sheet

# Dictionary to store daily idle times
daily_idle_times = {}

# Initial totals
day_total_production_time_in_seconds = 0
day_total_pt_seconds = 0
day_total_pt_trt_seconds = 0
previous_finish_time_in_seconds = None
previous_date = None

# Get shift times from the user
day_shift_start, day_shift_end, night_shift_start, night_shift_end = get_shift_time()

for part_report in root.findall('.//PartReports/PartReport'):
    part_name = part_report.find('PartName').text
    creation_datetime = part_report.find('TimeWhenPartWasCreated').text
    current_date = creation_datetime[:10]  # Here we extract the date from the datetime string

    time_object = datetime.strptime(creation_datetime, '%Y-%m-%dT%H:%M:%S')

    if current_date not in unique_dates:
        dashboard.append([current_date])  # This adds a new row with the date
        unique_dates[current_date] = True  # Mark this date as added

     
    if time_object.time() <= time(5, 0):  # Here we create a time object representing 5 AM
        # Adjust the date
        time_object -= timedelta(days=1)  # Here we subtract one day 
        
    current_date = time_object.strftime('%Y-%m-%d')  # Here we convert the time object back to a string    




    # Check if date changed and create a new sheet if needed
    if current_date != previous_date and previous_date is not None:
        # Append totals to the previous worksheet
        ws.append(["Totals", "", "", "",
                   seconds_to_decimal_hours(day_total_production_time_in_seconds),
                   seconds_to_decimal_hours(daily_idle_times.get(previous_date, 0)),
                   seconds_to_decimal_hours(day_total_pt_seconds),
                   seconds_to_decimal_hours(day_total_pt_trt_seconds)])

        adjust_column_width(ws)  # Auto-adjust columns' width

        #New Code that divides the production time total by 18hrs
        production_time_divided = seconds_to_decimal_hours(day_total_pt_seconds) / WORKING_TIME

        # Append this calculated value in a new row below the totals.
        ws.append(["Production Time / 24", "", "", "", "", "", production_time_divided, "", ""])



        # Reset totals for the new day
        day_total_production_time_in_seconds = 0
        day_total_pt_seconds = 0
        day_total_pt_trt_seconds = 0
        previous_finish_time_in_seconds = None

    if current_date != previous_date:
        ws = wb.create_sheet(title=current_date)
        # Adjusting the headers to include 'Shift' and indicate hours
        ws.append(["Part Name", "Date", "Start Time", "Finish Time", "Total Run Time (hours)", "Idle Time (hours)", 
                   "Production Time (hours)", "PT-TRT (hours)", "Shift"])

    start_time = creation_datetime[-8:]
    finish_datetime = part_report.find('TimeWhenPartWasFinished').text
    finish_time = finish_datetime[-8:]

    total_production_time_in_seconds = time_difference(start_time, finish_time)
    day_total_production_time_in_seconds += total_production_time_in_seconds

    production_time_seconds = time_to_seconds(part_report.find('TimeItTookToCreateThePart').text)
    day_total_pt_seconds += production_time_seconds

    pt_trt_in_seconds = total_production_time_in_seconds - production_time_seconds
    day_total_pt_trt_seconds += pt_trt_in_seconds  # This line was missing in the original script, causing an incorrect total PT-TRT.

   
    if previous_finish_time_in_seconds is not None:
        idle_time_in_seconds = time_to_seconds(start_time) - previous_finish_time_in_seconds
        if idle_time_in_seconds < 0:
            idle_time_in_seconds += 24 * 60 * 60  # Adjust for idle time that goes over midnight
    else:
        idle_time_in_seconds = 0

    daily_idle_times[current_date] = daily_idle_times.get(current_date, 0) + idle_time_in_seconds

    # Determine the shift for the current data row
    shift = determine_shift(start_time, day_shift_start, day_shift_end, night_shift_start, night_shift_end)

    # Prepare the row entry with time values converted to decimal hours.
    row_entry = [part_name, current_date, start_time, finish_time, 
                 seconds_to_decimal_hours(total_production_time_in_seconds),
                 seconds_to_decimal_hours(idle_time_in_seconds), 
                 seconds_to_decimal_hours(production_time_seconds), 
                 seconds_to_decimal_hours(pt_trt_in_seconds), shift]

    # Append the data to the master sheet as well as the current day's sheet
    master_sheet.append(row_entry)
    ws.append(row_entry)

    previous_date = current_date
    previous_finish_time_in_seconds = time_to_seconds(finish_time)

# Adjust the column width of the master sheet and other sheets
adjust_column_width(master_sheet)

# Append totals to the last worksheet
ws.append(["Totals", "", "", "",
           seconds_to_decimal_hours(day_total_production_time_in_seconds),
           seconds_to_decimal_hours(daily_idle_times.get(previous_date, 0)),
           seconds_to_decimal_hours(day_total_pt_seconds),
           seconds_to_decimal_hours(day_total_pt_trt_seconds)])

#New code 
production_time_divided = seconds_to_decimal_hours(day_total_pt_seconds) / WORKING_TIME

# Append this calculated value in a new row below the totals for the last date.
ws.append(["Production Time / 24", "", "", "", "", "", production_time_divided, "", ""])

# Remove the default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Save the workbook
output_excel_path = os.path.join(os.path.dirname(xml_file_path), "Production_Report_Beamline.xlsx")
wb.save(output_excel_path)

print(f"Excel file created successfully at {output_excel_path}!")