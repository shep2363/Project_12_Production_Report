import openpyxl
import os
from xml.etree import ElementTree as ET
from datetime import datetime, timedelta

# Constants and configurations
WORKING_TIME = 18  # Working hours
SHIFT_TIMES = {
    'day': ('06:00:00', '16:29:00'),
    'night': ('16:30:00', '03:00:00')
}

def parse_time(timestr):
    """Convert time string formatted as H:M:S to a datetime object."""
    return datetime.strptime(timestr, '%H:%M:%S')

def time_difference(time1, time2):
    """Calculate the difference between two times."""
    diff = (time2 - time1).seconds
    return diff if diff >= 0 else diff + 24 * 3600  # Adjust for time that goes over midnight

def get_shift(start_time):
    """Determine if a time falls within the day or night shift."""
    for shift, (start, end) in SHIFT_TIMES.items():
        if parse_time(start) <= start_time < parse_time(end):
            return shift.capitalize()
    return 'Undefined'  # For times that do not fall into either shift

def create_and_configure_workbook():
    """Create a new workbook with a pre-configured structure."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Create and set up the Dashboard sheet
    dashboard = wb.create_sheet(title="Dashboard", index=0)
    dashboard.append(["Date", "Production Efficiency"])

    return wb, dashboard

def main():
    xml_file_path = input("Enter the path to the XML file: ").strip('"')

    tree = ET.parse(xml_file_path)
    root = tree.getroot()

    wb, dashboard = create_and_configure_workbook()
    master_sheet = wb.create_sheet(title="Master Sheet")
    master_sheet_headers = ["Part Name", "Date", "Start Time", "Finish Time", "Total Run Time (hours)",
                            "Idle Time (hours)", "Production Time (hours)", "PT-TRT (hours)", "Shift"]
    master_sheet.append(master_sheet_headers)

    unique_dates = {}
    current_sheet = None

    for part_report in root.findall('.//PartReports/PartReport'):
        part_name = part_report.find('PartName').text
        start_time = part_report.find('TimeWhenPartWasCreated').text
        finish_time = part_report.find('TimeWhenPartWasFinished').text

        # Convert times to datetime objects
        start_time_obj = datetime.strptime(start_time, "%Y-%m-%dT%H:%M:%S")
        finish_time_obj = datetime.strptime(finish_time, "%Y-%m-%dT%H:%M:%S")

        # Extract the date (without time) from the datetime object
        current_date = start_time_obj.date()

        total_run_time_seconds = time_difference(start_time_obj, finish_time_obj)
        total_run_time_hours = total_run_time_seconds / 3600  # Convert seconds to hours

        # Assume a predefined working time for calculations
        production_time_hours = WORKING_TIME
        idle_time_hours = max(0, total_run_time_hours - production_time_hours)
        pt_trt_hours = production_time_hours - total_run_time_hours

        current_date_str = current_date.strftime('%Y-%m-%d')

        if current_date_str not in unique_dates:
            unique_dates[current_date_str] = True
            dashboard.append([current_date_str])
            current_sheet = wb.create_sheet(title=current_date_str)
            current_sheet.append(master_sheet_headers)

        shift = get_shift(start_time_obj.time())

        row_data = [part_name, current_date_str, start_time, finish_time, total_run_time_hours,
                    idle_time_hours, production_time_hours, pt_trt_hours, shift]

        master_sheet.append(row_data)
        current_sheet.append(row_data)

    output_excel_path = os.path.join(os.path.dirname(xml_file_path), "Production_Report.xlsx")
    wb.save(output_excel_path)
    print(f"Excel file created successfully at {output_excel_path}!")

if __name__ == "__main__":
    main()
